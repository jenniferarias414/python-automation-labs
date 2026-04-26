from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path


BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"

INPUT_FILE = DATA_DIR / "data.xlsx"
OUTPUT_FILE = OUTPUT_DIR / "report.xlsx"


def create_sample_excel_file():
    """Create sample Excel data if the input file does not already exist."""

    DATA_DIR.mkdir(exist_ok=True)

    if INPUT_FILE.exists():
        return

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"

    rows = [
        ["Month", "Sales", "Costs", "Profit", "Units"],
        ["January", 1000, 450, 550, 25],
        ["February", 1500, 700, 800, 32],
        ["March", 1250, 500, 750, 28],
        ["April", 1750, 850, 900, 40],
    ]

    for row in rows:
        worksheet.append(row)

    workbook.save(INPUT_FILE)


def generate_monthly_report(input_file, output_file):
    """Read Excel data, calculate totals, and create a formatted report sheet."""

    OUTPUT_DIR.mkdir(exist_ok=True)

    workbook = load_workbook(input_file)
    data_sheet = workbook["Data"]

    totals = {}

    # Columns B through E contain numeric values.
    for column in range(2, 6):
        header = data_sheet.cell(row=1, column=column).value
        total = 0

        for row in range(2, data_sheet.max_row + 1):
            value = data_sheet.cell(row=row, column=column).value

            if isinstance(value, (int, float)):
                total += value

        totals[header] = total

    if "Monthly Report" in workbook.sheetnames:
        del workbook["Monthly Report"]

    report_sheet = workbook.create_sheet(title="Monthly Report")

    report_sheet["A1"] = "Monthly Totals Report"
    report_sheet["A1"].font = Font(bold=True, size=14)
    report_sheet["A1"].fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    report_sheet.merge_cells("A1:B1")

    report_sheet["A3"] = "Metric"
    report_sheet["B3"] = "Total"

    for cell in report_sheet[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    row_number = 4

    for metric, total in totals.items():
        report_sheet.cell(row=row_number, column=1).value = metric
        report_sheet.cell(row=row_number, column=2).value = total
        row_number += 1

    report_sheet.freeze_panes = "A4"

    for column_cells in report_sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        report_sheet.column_dimensions[column_letter].width = 20

    workbook.save(output_file)

    print(f"Monthly report created: {output_file}")


if __name__ == "__main__":
    create_sample_excel_file()
    generate_monthly_report(INPUT_FILE, OUTPUT_FILE)